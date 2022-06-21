from openpyxl import load_workbook

class pyxl_handler:
    def __init__(self, xl):
        self.excel = load_workbook(xl)
        # self.excel.save(xl)
        pass

    def get_cell(self, sheet, x, y):
        data = self.excel[sheet].cell(x,y).value
        return data

    def get_table(self, sheet):
        m_row = self.excel[sheet].max_row
        m_col = self.excel[sheet].max_column
        tab_lst = []
        for i in range(1, m_row + 1):
            inn_lst = []
            for ii in range(1, m_col + 1):
                inn_lst.append(self.excel[sheet].cell(i,ii).value)
            tab_lst.append(inn_lst)
        return tab_lst